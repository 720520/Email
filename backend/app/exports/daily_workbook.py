"""每日基金净值汇总工作簿构建。"""

from __future__ import annotations

import re
from collections.abc import Sequence
from datetime import date, datetime
from decimal import Decimal

from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.table import Table, TableStyleInfo

from app.exports.models import DailyNavExportRow, ExceptionExportRow

_ILLEGAL_XML_CHARACTERS = re.compile(r"[\x00-\x08\x0B-\x0C\x0E-\x1F]")
_FORMULA_PREFIXES = ("=", "+", "-", "@", "\t", "\r")

_DARK_BLUE = "17365D"
_MEDIUM_BLUE = "2F75B5"
_LIGHT_BLUE = "D9EAF7"
_LIGHT_BORDER = "D9E2F3"
_ERROR_FILL = "FDE9E7"
_ERROR_FONT = "9C0006"
_WARNING_FILL = "FFF2CC"
_WARNING_FONT = "7F6000"


class DailyNavWorkbookBuilder:
    """构建便于运营复核、筛选和打印的双工作表报表。"""

    title_row = 1
    metadata_row = 2
    summary_row = 3
    header_row = 5
    first_data_row = 6

    def build(
        self,
        *,
        report_date: date,
        generated_at: datetime,
        nav_rows: Sequence[DailyNavExportRow],
        exception_rows: Sequence[ExceptionExportRow],
    ) -> Workbook:
        workbook = Workbook()
        nav_sheet = workbook.active
        nav_sheet.title = "基金净值"
        exception_sheet = workbook.create_sheet("异常记录")

        self._build_nav_sheet(
            nav_sheet,
            report_date=report_date,
            generated_at=generated_at,
            rows=nav_rows,
            exception_count=len(exception_rows),
        )
        self._build_exception_sheet(
            exception_sheet,
            report_date=report_date,
            generated_at=generated_at,
            rows=exception_rows,
        )
        workbook.properties.creator = "基金运营邮件自动解析与净值汇总系统"
        workbook.properties.title = f"{report_date.isoformat()} 每日基金净值汇总"
        workbook.properties.subject = "基金净值与运营异常审计"
        workbook.calculation.fullCalcOnLoad = True
        workbook.calculation.forceFullCalc = True
        workbook.calculation.calcMode = "auto"
        return workbook

    def _build_nav_sheet(
        self,
        sheet,
        *,
        report_date: date,
        generated_at: datetime,
        rows: Sequence[DailyNavExportRow],
        exception_count: int,
    ) -> None:
        headers = ["日期", "产品代码", "产品名称", "单位净值", "累计净值", "资产净值", "来源"]
        last_row = self.header_row + len(rows)
        formula_last_row = max(self.first_data_row, last_row)

        sheet.merge_cells("A1:G1")
        sheet["A1"] = "每日基金净值汇总"
        sheet["A2"] = "净值日期"
        sheet["B2"] = report_date
        sheet["C2"] = "生成时间"
        sheet["D2"] = _excel_datetime(generated_at)
        sheet["E2"] = "数据口径"
        sheet.merge_cells("F2:G2")
        sheet["F2"] = "产品代码 + 日期去重后的有效记录"
        sheet["A3"] = "基金数量"
        sheet["B3"] = f"=COUNTA(B{self.first_data_row}:B{formula_last_row})"
        sheet["C3"] = "异常数量"
        sheet["D3"] = exception_count
        sheet["E3"] = "资产净值合计"
        sheet["F3"] = f"=SUM(F{self.first_data_row}:F{formula_last_row})"
        sheet.merge_cells("F3:G3")
        sheet.append([])
        sheet.append(headers)

        for row in rows:
            sheet.append(
                [
                    row.nav_date,
                    _safe_text(row.product_code),
                    _safe_text(row.product_name),
                    _number(row.unit_nav),
                    _number(row.total_nav),
                    _number(row.asset_value),
                    _safe_text(row.source_file),
                ]
            )

        self._style_common_sheet(sheet, last_column=7)
        self._style_title(sheet, "A1:G1")
        self._style_metadata(sheet, "A2:G2")
        self._style_summary(sheet, "A3:G3")
        sheet["B3"].number_format = "#,##0"
        sheet["D3"].number_format = "#,##0"
        sheet["F3"].number_format = "#,##0.00;[Red](#,##0.00);-"
        self._style_header(sheet, "A5:G5")
        self._style_nav_data(sheet, last_row)
        self._add_table_or_filter(sheet, "A5", "G", last_row, "FundNavTable")
        self._set_column_widths(
            sheet,
            {"A": 13, "B": 17, "C": 38, "D": 16, "E": 16, "F": 22, "G": 42},
        )
        sheet.freeze_panes = "A6"
        sheet.auto_filter.ref = f"A5:G{max(self.header_row, last_row)}"
        sheet.sheet_properties.pageSetUpPr.fitToPage = True
        sheet.page_setup.orientation = "landscape"
        sheet.page_setup.fitToWidth = 1
        sheet.page_setup.fitToHeight = 0
        sheet.print_title_rows = "1:5"
        sheet.print_area = f"A1:G{max(self.header_row, last_row)}"

    def _build_exception_sheet(
        self,
        sheet,
        *,
        report_date: date,
        generated_at: datetime,
        rows: Sequence[ExceptionExportRow],
    ) -> None:
        headers = [
            "日期",
            "异常分类",
            "严重级别",
            "产品代码",
            "产品名称",
            "来源",
            "工作表",
            "行号",
            "字段",
            "原始值",
            "异常说明",
            "处理状态",
        ]
        last_row = self.header_row + len(rows)
        formula_last_row = max(self.first_data_row, last_row)

        sheet.merge_cells("A1:L1")
        sheet["A1"] = "异常记录"
        sheet["A2"] = "统计日期"
        sheet["B2"] = report_date
        sheet["C2"] = "生成时间"
        sheet["D2"] = _excel_datetime(generated_at)
        sheet["E2"] = "说明"
        sheet.merge_cells("F2:L2")
        sheet["F2"] = "异常按创建日期统计；历史邮件重新解析产生的异常计入重解析当日"
        sheet["A3"] = "异常总数"
        sheet["B3"] = f"=COUNTA(B{self.first_data_row}:B{formula_last_row})"
        sheet["C3"] = "错误"
        sheet["D3"] = (
            f'=COUNTIF(C{self.first_data_row}:C{formula_last_row},"错误")'
        )
        sheet["E3"] = "警告"
        sheet["F3"] = (
            f'=COUNTIF(C{self.first_data_row}:C{formula_last_row},"警告")'
        )
        sheet.append([])
        sheet.append(headers)

        for row in rows:
            sheet.append(
                [
                    row.occurred_date,
                    _safe_text(row.category),
                    _safe_text(row.severity),
                    _safe_text(row.product_code),
                    _safe_text(row.product_name),
                    _safe_text(row.source),
                    _safe_text(row.sheet_name),
                    row.row_number,
                    _safe_text(row.field_name),
                    _safe_text(row.raw_value),
                    _safe_text(row.message),
                    _safe_text(row.status),
                ]
            )

        self._style_common_sheet(sheet, last_column=12)
        self._style_title(sheet, "A1:L1")
        self._style_metadata(sheet, "A2:L2")
        self._style_summary(sheet, "A3:L3")
        for cell_reference in ("B3", "D3", "F3"):
            sheet[cell_reference].number_format = "#,##0"
        self._style_header(sheet, "A5:L5")
        self._style_exception_data(sheet, last_row)
        self._add_table_or_filter(sheet, "A5", "L", last_row, "ExceptionTable")
        self._set_column_widths(
            sheet,
            {
                "A": 13,
                "B": 16,
                "C": 11,
                "D": 17,
                "E": 30,
                "F": 38,
                "G": 18,
                "H": 9,
                "I": 18,
                "J": 22,
                "K": 48,
                "L": 12,
            },
        )
        sheet.freeze_panes = "A6"
        sheet.auto_filter.ref = f"A5:L{max(self.header_row, last_row)}"
        if rows:
            severity_range = f"C{self.first_data_row}:C{last_row}"
            sheet.conditional_formatting.add(
                severity_range,
                FormulaRule(
                    formula=[f'C{self.first_data_row}="错误"'],
                    fill=PatternFill("solid", fgColor=_ERROR_FILL),
                    font=Font(color=_ERROR_FONT),
                ),
            )
            sheet.conditional_formatting.add(
                severity_range,
                FormulaRule(
                    formula=[f'C{self.first_data_row}="警告"'],
                    fill=PatternFill("solid", fgColor=_WARNING_FILL),
                    font=Font(color=_WARNING_FONT),
                ),
            )
        sheet.sheet_properties.pageSetUpPr.fitToPage = True
        sheet.page_setup.orientation = "landscape"
        sheet.page_setup.fitToWidth = 1
        sheet.page_setup.fitToHeight = 0
        sheet.print_title_rows = "1:5"
        sheet.print_area = f"A1:L{max(self.header_row, last_row)}"

    @staticmethod
    def _style_common_sheet(sheet, *, last_column: int) -> None:
        sheet.sheet_view.showGridLines = False
        sheet.sheet_view.zoomScale = 90
        sheet.row_dimensions[1].height = 32
        sheet.row_dimensions[2].height = 24
        sheet.row_dimensions[3].height = 26
        sheet.row_dimensions[4].height = 9
        sheet.row_dimensions[5].height = 27
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, max_col=last_column):
            for cell in row:
                cell.font = Font(name="等线", size=10, color="1F1F1F")
                cell.alignment = Alignment(vertical="center")

    @staticmethod
    def _style_title(sheet, cell_range: str) -> None:
        cells = sheet[cell_range]
        for row in cells:
            for cell in row:
                cell.fill = PatternFill("solid", fgColor=_DARK_BLUE)
                cell.font = Font(name="等线", size=16, bold=True, color="FFFFFF")
                cell.alignment = Alignment(horizontal="left", vertical="center")

    @staticmethod
    def _style_metadata(sheet, cell_range: str) -> None:
        label_columns = {1, 3, 5}
        for row in sheet[cell_range]:
            for cell in row:
                cell.fill = PatternFill("solid", fgColor="F4F7FB")
                if cell.column in label_columns:
                    cell.font = Font(name="等线", size=10, bold=True, color=_DARK_BLUE)
        sheet["B2"].number_format = "yyyy-mm-dd"
        sheet["D2"].number_format = "yyyy-mm-dd hh:mm:ss"

    @staticmethod
    def _style_summary(sheet, cell_range: str) -> None:
        label_columns = {1, 3, 5}
        for row in sheet[cell_range]:
            for cell in row:
                cell.fill = PatternFill("solid", fgColor=_LIGHT_BLUE)
                cell.border = Border(bottom=Side(style="thin", color=_MEDIUM_BLUE))
                if cell.column in label_columns:
                    cell.font = Font(name="等线", size=10, bold=True, color=_DARK_BLUE)
                else:
                    cell.font = Font(name="等线", size=11, bold=True, color="000000")
    @staticmethod
    def _style_header(sheet, cell_range: str) -> None:
        for row in sheet[cell_range]:
            for cell in row:
                cell.fill = PatternFill("solid", fgColor=_MEDIUM_BLUE)
                cell.font = Font(name="等线", size=10, bold=True, color="FFFFFF")
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = Border(bottom=Side(style="medium", color=_DARK_BLUE))

    def _style_nav_data(self, sheet, last_row: int) -> None:
        if last_row < self.first_data_row:
            return
        for row_number in range(self.first_data_row, last_row + 1):
            sheet.row_dimensions[row_number].height = 32
            sheet[f"A{row_number}"].number_format = "yyyy-mm-dd"
            sheet[f"B{row_number}"].number_format = "@"
            sheet[f"D{row_number}"].number_format = "0.00000000;[Red](0.00000000);-"
            sheet[f"E{row_number}"].number_format = "0.00000000;[Red](0.00000000);-"
            sheet[f"F{row_number}"].number_format = "#,##0.0000;[Red](#,##0.0000);-"
            for column in (4, 5, 6):
                sheet.cell(row=row_number, column=column).alignment = Alignment(
                    horizontal="right", vertical="center"
                )
            sheet.cell(row=row_number, column=3).alignment = Alignment(
                horizontal="left", vertical="center", wrap_text=False
            )
            sheet.cell(row=row_number, column=7).alignment = Alignment(
                horizontal="left", vertical="center", wrap_text=True
            )
            self._add_row_separator(sheet, row_number, last_column=7)

    def _style_exception_data(self, sheet, last_row: int) -> None:
        if last_row < self.first_data_row:
            return
        for row_number in range(self.first_data_row, last_row + 1):
            sheet.row_dimensions[row_number].height = 32
            sheet[f"A{row_number}"].number_format = "yyyy-mm-dd"
            sheet[f"D{row_number}"].number_format = "@"
            sheet[f"H{row_number}"].number_format = "0"
            for column in (5, 6, 10, 11):
                sheet.cell(row=row_number, column=column).alignment = Alignment(
                    horizontal="left", vertical="center", wrap_text=True
                )
            self._add_row_separator(sheet, row_number, last_column=12)

    @staticmethod
    def _add_row_separator(sheet, row_number: int, *, last_column: int) -> None:
        border = Border(bottom=Side(style="hair", color=_LIGHT_BORDER))
        for column in range(1, last_column + 1):
            sheet.cell(row=row_number, column=column).border = border

    @staticmethod
    def _add_table_or_filter(
        sheet,
        start_cell: str,
        last_column: str,
        last_row: int,
        table_name: str,
    ) -> None:
        if last_row <= DailyNavWorkbookBuilder.header_row:
            sheet.auto_filter.ref = (
                f"{start_cell}:{last_column}{DailyNavWorkbookBuilder.header_row}"
            )
            return
        table = Table(
            displayName=table_name,
            ref=f"{start_cell}:{last_column}{last_row}",
        )
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        sheet.add_table(table)

    @staticmethod
    def _set_column_widths(sheet, widths: dict[str, float]) -> None:
        for column, width in widths.items():
            sheet.column_dimensions[column].width = width


def _number(value: Decimal | None) -> Decimal | None:
    return value


def _excel_datetime(value: datetime) -> datetime:
    return value.replace(tzinfo=None) if value.tzinfo is not None else value


def _safe_text(value: object | None, *, max_length: int = 32767) -> str | None:
    if value is None:
        return None
    text = _ILLEGAL_XML_CHARACTERS.sub("", str(value))[:max_length]
    if text.startswith(_FORMULA_PREFIXES):
        return f"'{text}"
    return text
