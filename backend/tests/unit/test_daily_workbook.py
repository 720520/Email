from __future__ import annotations

from datetime import UTC, date, datetime
from decimal import Decimal
from pathlib import Path

from openpyxl import load_workbook

from app.exports import DailyNavExportRow, DailyNavWorkbookBuilder, ExceptionExportRow


def test_daily_workbook_has_required_sheets_values_and_controls(tmp_path: Path) -> None:
    builder = DailyNavWorkbookBuilder()
    workbook = builder.build(
        report_date=date(2026, 7, 24),
        generated_at=datetime(2026, 7, 24, 18, tzinfo=UTC),
        nav_rows=[
            DailyNavExportRow(
                nav_date=date(2026, 7, 24),
                product_code="SAWK26",
                product_name="吉余宸锋金炜幸福一号私募证券投资基金",
                unit_nav=Decimal("1.23456789"),
                total_nav=Decimal("1.34567891"),
                asset_value=Decimal("123456789.1234"),
                source_file="净值表.xlsx",
            )
        ],
        exception_rows=[
            ExceptionExportRow(
                occurred_date=date(2026, 7, 24),
                category="产品重复",
                severity="错误",
                product_code="SAWK26",
                product_name="吉余宸锋金炜幸福一号私募证券投资基金",
                source="重复净值.xlsx",
                sheet_name="基金净值",
                row_number=8,
                field_name="product_code+nav_date",
                raw_value="=HYPERLINK(\"bad\")",
                message="产品代码和日期已存在，已保留历史记录并拒绝覆盖",
                status="待处理",
            )
        ],
    )
    output_path = tmp_path / "每日基金净值汇总.xlsx"
    workbook.save(output_path)
    workbook.close()

    loaded = load_workbook(output_path, data_only=False)
    assert loaded.sheetnames == ["基金净值", "异常记录"]

    nav_sheet = loaded["基金净值"]
    assert nav_sheet["A1"].value == "每日基金净值汇总"
    assert nav_sheet["A5"].value == "日期"
    assert nav_sheet["G5"].value == "来源"
    assert nav_sheet["B6"].value == "SAWK26"
    assert nav_sheet["D6"].value == 1.23456789
    assert nav_sheet["F6"].value == 123456789.1234
    assert nav_sheet["B3"].value == "=COUNTA(B6:B6)"
    assert nav_sheet.freeze_panes == "A6"
    assert nav_sheet.sheet_view.showGridLines is False
    assert set(nav_sheet.tables) == {"FundNavTable"}
    assert nav_sheet["D6"].number_format == "0.00000000;[Red](0.00000000);-"
    assert nav_sheet["F6"].number_format == "#,##0.0000;[Red](#,##0.0000);-"

    exception_sheet = loaded["异常记录"]
    assert exception_sheet["B6"].value == "产品重复"
    assert exception_sheet["C6"].value == "错误"
    assert exception_sheet["F3"].number_format == "#,##0"
    assert exception_sheet["J6"].data_type == "s"
    assert exception_sheet["J6"].value.startswith("'")
    assert exception_sheet.freeze_panes == "A6"
    assert set(exception_sheet.tables) == {"ExceptionTable"}
    conditional_ranges = list(exception_sheet.conditional_formatting)
    assert len(conditional_ranges) == 1
    assert len(exception_sheet.conditional_formatting[conditional_ranges[0]]) == 2
    loaded.close()


def test_empty_daily_workbook_still_has_headers_and_filters(tmp_path: Path) -> None:
    workbook = DailyNavWorkbookBuilder().build(
        report_date=date(2026, 7, 24),
        generated_at=datetime(2026, 7, 24, 18, tzinfo=UTC),
        nav_rows=[],
        exception_rows=[],
    )
    output_path = tmp_path / "empty.xlsx"
    workbook.save(output_path)
    workbook.close()

    loaded = load_workbook(output_path, data_only=False)
    assert loaded["基金净值"].auto_filter.ref == "A5:G5"
    assert loaded["异常记录"].auto_filter.ref == "A5:L5"
    assert loaded["基金净值"]["B3"].value == "=COUNTA(B6:B6)"
    assert loaded["异常记录"]["B3"].value == "=COUNTA(B6:B6)"
    loaded.close()
